import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
import requests
import pandas as pd
from datetime import datetime, timedelta
import openpyxl

app = Flask(__name__)
app.secret_key = "radar_robustec_v2_ultra"

# Configurações do Supabase
SUPABASE_URL = "https://ndfpegcuwthqsjpayfmp.supabase.co"
SUPABASE_KEY = "sb_publishable_wLoOr_q9VoTj7V7VQJLgVw_oS5Fycke"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

EXCEL_FILE = "ACOMPANHAMENTO PROCESSUAL 2023.xlsx"

def get_week_range():
    today = datetime.now()
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=4)
    return f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

def fetch_all_processes():
    """Busca todos os processos usando paginação para superar o limite de 1000."""
    all_data = []
    limit = 1000
    offset = 0
    while True:
        headers_req = HEADERS.copy()
        headers_req["Range"] = f"{offset}-{offset + limit - 1}"
        response = requests.get(f"{SUPABASE_URL}/rest/v1/processos?select=*&order=created_at.desc", headers=headers_req)
        if response.status_code not in [200, 206] or not response.json():
            break
        data = response.json()
        all_data.extend(data)
        if len(data) < limit:
            break
        offset += limit
    return all_data

def update_excel_local(sinistro_or_numero, andamento, responsavel):
    """Atualiza o arquivo Excel local na aba do responsável."""
    if not os.path.exists(EXCEL_FILE):
        return False
    
    try:
        # Tentar abrir com openpyxl para preservar formatação se possível, 
        # ou usar pandas para simplicidade se não houver fórmulas complexas.
        xl = pd.ExcelFile(EXCEL_FILE)
        
        # O usuário disse que o processo 2017183928 era da Marcia
        # Precisamos descobrir em qual aba o processo está se o 'responsavel' não for exato
        target_sheet = responsavel
        if target_sheet not in xl.sheet_names:
            # Busca em todas as abas
            for sheet in xl.sheet_names:
                df_temp = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
                mask = (df_temp.astype(str).apply(lambda x: x.str.contains(sinistro_or_numero)).any(axis=1))
                if mask.any():
                    target_sheet = sheet
                    break
        
        # Abrir o workbook para escrita
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
            # Encontrar a linha (busca na coluna de Número ou Sinistro)
            found = False
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if any(str(val) == str(sinistro_or_numero) for val in row if val):
                    # Encontrou a linha. Agora encontra a última coluna disponível
                    last_col = ws.max_column
                    # O usuário quer na "última coluna da linha"
                    # Vamos adicionar em uma nova coluna se a última estiver ocupada
                    ws.cell(row=row_idx, column=last_col + 1, value=f"{datetime.now().strftime('%d/%m/%Y')}: {andamento}")
                    found = True
                    break
            
            if found:
                wb.save(EXCEL_FILE)
                return True
    except Exception as e:
        print(f"Erro ao atualizar Excel: {e}")
    return False

@app.route("/")
def index():
    processos = fetch_all_processes()
    
    # Estatísticas
    total = len(processos)
    ativos = sum(1 for p in processos if p.get("status") != "Encerrado")
    encerrados = sum(1 for p in processos if p.get("status") == "Encerrado")
    analistas = list(set(p.get("responsavel") for p in processos if p.get("responsavel")))
    
    estatisticas = {
        "total": total,
        "ativos": ativos,
        "encerrados": encerrados,
        "semana": get_week_range(),
        "analistas": sorted(analistas)
    }

    # Gráficos com normalização
    contagem_analista = {}
    contagem_sistema = {}
    for p in processos:
        r = p.get('responsavel')
        s = p.get('sistema')
        if r: contagem_analista[r] = contagem_analista.get(r, 0) + 1
        
        # Normalização de Sistemas
        if s:
            s_up = str(s).upper().strip()
            if 'SAJ' in s_up: s = 'E-SAJ'
            elif 'PJE' in s_up: s = 'PJE'
            elif 'PROC' in s_up: s = 'E-PROC'
            elif 'ELETRONICO' in s_up: s = 'ELETRÔNICO'
            else: s = s_up
            contagem_sistema[s] = contagem_sistema.get(s, 0) + 1
    
    # Ordem fixa solicitada pelo usuário
    ordem_analistas = ["Miguel_1", "Carmem", "Caroline", "Victor", "Miguel", "Carolina", "Marcia", "Debora"]
    
    # Gráfico de Analistas: Voltar para a ordem decrescente por quantidade ("em ordem como estava")
    analistas_grafico = [{"name": r, "y": c} for r, c in sorted(contagem_analista.items(), key=lambda x: x[1], reverse=True)]
    
    sistemas_grafico = [{"name": s, "y": c} for s, c in sorted(contagem_sistema.items(), key=lambda x: x[1], reverse=True)]

    return render_template("index.html", 
                           processos=processos, 
                           estatisticas=estatisticas, 
                           analistas_grafico=analistas_grafico,
                           sistemas_grafico=sistemas_grafico,
                           ordem_analistas=ordem_analistas)

@app.route("/api/processos", methods=["POST"])
def novo_processo():
    dados = {
        "numero_processo": request.form.get("numero_processo"),
        "sinistro_allianz": request.form.get("sinistro_allianz"),
        "autor": request.form.get("autor"),
        "comarca": request.form.get("comarca"),
        "uf": request.form.get("uf"),
        "sistema": request.form.get("sistema"),
        "responsavel": request.form.get("responsavel"),
        "status": "Ativo"
    }
    response = requests.post(f"{SUPABASE_URL}/rest/v1/processos", headers=HEADERS, json=dados)
    if response.status_code in [200, 201]:
        flash("Processo criado com sucesso!", "success")
    else:
        flash("Erro ao criar processo.", "error")
    return redirect(url_for("index"))

@app.route("/api/busca_processo")
def busca_processo():
    query = request.args.get("q", "")
    if not query: return jsonify([])
    # Busca por Numero ou Sinistro e retorna o responsavel
    url = f"{SUPABASE_URL}/rest/v1/processos?or=(numero_processo.eq.{query},sinistro_allianz.eq.{query})&select=id,numero_processo,sinistro_allianz,autor,responsavel"
    response = requests.get(url, headers=HEADERS)
    return jsonify(response.json() if response.status_code == 200 else [])

@app.route("/api/andamentos", methods=["POST"])
def novo_andamento():
    processo_id = request.form.get("processo_id")
    descricao = request.form.get("descricao")
    responsavel = request.form.get("responsavel_nome")
    identificador = request.form.get("identificador") # Numero ou Sinistro para o Excel

    # 1. Supabase
    requests.patch(f"{SUPABASE_URL}/rest/v1/processos?id=eq.{processo_id}", headers=HEADERS, json={
        "ultimo_andamento": descricao,
        "data_atualizacao": datetime.now().isoformat()
    })

    # 2. Excel Local (Sincronização opcional)
    if identificador and responsavel:
        success = update_excel_local(identificador, descricao, responsavel)
        if success:
            flash("Andamento salvo no Supabase e no Excel!", "success")
        else:
            # Em produção (Render), não mostramos erro se o Excel falhar, pois o Supabase é a fonte da verdade
            if os.environ.get("RENDER"):
                flash("Andamento salvo no Supabase!", "success")
            else:
                flash("Andamento salvo no Supabase, mas houve erro ao atualizar o Excel local.", "warning")
    else:
        flash("Andamento salvo no Supabase.", "success")
    
    return redirect(url_for("index"))

@app.route("/api/andamentos/editar", methods=["POST"])
def editar_andamento():
    processo_id = request.form.get("processo_id")
    nova_descricao = request.form.get("descricao")
    
    # 1. Atualizar Supabase (último_andamento)
    response = requests.patch(f"{SUPABASE_URL}/rest/v1/processos?id=eq.{processo_id}", headers=HEADERS, json={
        "ultimo_andamento": nova_descricao,
        "data_atualizacao": datetime.now().isoformat()
    })
    
    if response.status_code in [200, 204]:
        flash("Andamento editado com sucesso!", "success")
    else:
        flash("Erro ao editar andamento.", "error")
        
    return redirect(url_for("index"))

@app.route("/api/processos/encerrar", methods=["POST"])
def encerrar_processo():
    processo_id = request.form.get("processo_id")
    
    response = requests.patch(f"{SUPABASE_URL}/rest/v1/processos?id=eq.{processo_id}", headers=HEADERS, json={
        "status": "Encerrado"
    })
    
    if response.status_code in [200, 204]:
        flash("Processo encerrado com sucesso!", "success")
    else:
        flash("Erro ao encerrar processo.", "error")
        
    return redirect(url_for("index"))

@app.route("/api/exportar")
def exportar_excel():
    processos = fetch_all_processes()
    df = pd.DataFrame(processos)
    
    # Limpeza básica para o Excel
    if not df.empty:
        colunas_remover = ['id', 'created_at']
        df = df.drop(columns=[c for c in colunas_remover if c in df.columns])
    
    output_file = f"EXPORT_SISTEMA_BRUNIERA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    df.to_excel(output_file, index=False)
    
    return jsonify({"success": True, "file": output_file})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
